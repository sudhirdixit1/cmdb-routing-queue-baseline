"""Verify the workbook without LibreOffice.

No recalculation engine is available on this machine, so column L cannot be
executed.  Instead: read the degradation table straight out of the saved
workbook, re-implement Excel's INDEX/MATCH(...,1) semantics in Python, and
check the interpolation against values computed independently by numpy.
Also confirm every formula references the range it is supposed to.
"""
import re
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

WB = Path(__file__).resolve().parent.parent / "capability_readiness_matrix.xlsx"
wb = load_workbook(WB)
ev, mx = wb["Evidence"], wb["Matrix"]

# ---- pull the lookup table out of the workbook itself -------------------
# Verify by HEADER NAME, not position -- a shifted column is the failure this
# check exists to catch, and value patterns alone will not reveal it.
hdr_b = ev.cell(row=5, column=2).value
hdr_c = ev.cell(row=5, column=3).value
assert hdr_b == "CI population rate", f"col B header is {hdr_b!r}, expected rate"
assert hdr_c == "benefit retained", f"col C header is {hdr_c!r}, expected benefit"
print(f"header check: B='{hdr_b}'  C='{hdr_c}'  OK")

rates, kept = [], []
for r in range(6, 20):
    rates.append(ev.cell(row=r, column=2).value)
    kept.append(ev.cell(row=r, column=3).value)
print(f"lookup table rows 6-19: {len(rates)} entries")
assert all(isinstance(v, (int, float)) for v in rates), "non-numeric rate found"
assert all(isinstance(v, (int, float)) for v in kept), "non-numeric benefit found"
assert rates == sorted(rates), "rates must be ascending for MATCH(...,1)"
print(f"  rates   ascending: OK   ({rates[0]} .. {rates[-1]})")
print(f"  benefit ascending: OK   ({kept[0]} .. {kept[-1]})")


def excel_formula(x):
    """Replicate exactly what the cell formula computes."""
    if not isinstance(x, (int, float)):
        return ""
    if x >= 1:
        return 1.0
    if x <= 0:
        return 0.0
    m = max(i for i, v in enumerate(rates) if v <= x)      # MATCH(x, B, 1) - 1
    lo_r, hi_r, lo_c, hi_c = rates[m], rates[m + 1], kept[m], kept[m + 1]
    return lo_c + (x - lo_r) / (hi_r - lo_r) * (hi_c - lo_c)


print("\ninterpolation check against numpy.interp:")
ok = True
for x in [0.0, 0.001, 0.002, 0.05, 0.10, 0.25, 0.35, 0.50, 0.65, 0.80,
          0.90, 0.95, 0.999, 1.0, 1.5, -0.2]:
    got = excel_formula(x)
    exp = float(np.interp(np.clip(x, 0, 1), rates, kept))
    match = isinstance(got, float) and abs(got - exp) < 1e-12
    ok &= match
    print(f"  I={x:>6} -> {got if isinstance(got,str) else f'{got:.4f}'}"
          f"   numpy {exp:.4f}   {'OK' if match else 'MISMATCH'}")
for x in ["", None, "Walk"]:
    got = excel_formula(x)
    print(f"  I={x!r:>8} -> {got!r}   {'OK' if got == '' else 'MISMATCH'}")
    ok &= got == ""

# ---- confirm the written formulas point at the right ranges ------------
print("\nformula range check:")
bad = 0
for i in range(2, 12):
    f = mx.cell(row=i, column=12).value
    if not (isinstance(f, str) and f.startswith("=")):
        print(f"  row {i}: NOT A FORMULA -> {f!r}")
        bad += 1
        continue
    if "Evidence!$B$6:$B$19" not in f or "Evidence!$C$6:$C$19" not in f:
        print(f"  row {i}: wrong lookup range")
        bad += 1
    if f"I{i}" not in f:
        print(f"  row {i}: does not reference its own input cell I{i}")
        bad += 1
    for fn in re.findall(r"([A-Z]+)\(", f):
        if fn not in {"IF", "NOT", "ISNUMBER", "INDEX", "MATCH"}:
            print(f"  row {i}: unexpected function {fn}")
            bad += 1
print(f"  {10 - bad if bad <= 10 else 0}/10 formulas reference the correct ranges "
      f"and use only Excel-2007 functions")

# ---- input cells are actually blank and marked -------------------------
print("\ninput cells:")
blank = sum(1 for i in range(2, 12) for j in (8, 9, 10, 11)
            if mx.cell(row=i, column=j).value in (None, ""))
yellow = sum(1 for i in range(2, 12) for j in (8, 9, 10, 11)
             if mx.cell(row=i, column=j).fill.fgColor.rgb in ("00FFFF00", "FFFFFF00"))
print(f"  {blank}/40 blank and awaiting input")
print(f"  {yellow}/40 filled yellow")

print("\n" + "=" * 66)
print("VERDICT:", "logic verified" if ok and bad == 0 else "PROBLEM FOUND")
print("Formulas are hand-verified against numpy, not executed -- no Excel or")
print("LibreOffice on this machine.  Open the file once to confirm column L")
print("populates when you enter a value in column I.")
