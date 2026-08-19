"""Build the capability-readiness matrix workbook for practitioner completion.

Data-derived columns are pre-filled from the measured results.  Cells the
practitioner must supply are filled yellow.  Column L computes, by formula,
the modelled benefit retained at whatever CI population rate they enter --
so their domain judgement produces a quantitative result rather than a label.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent.parent / "capability_readiness_matrix.xlsx"

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="0D6B6E")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
SECT_FILL = PatternFill("solid", fgColor="E8EEF1")
CALC_FILL = PatternFill("solid", fgColor="E2F0F0")
THIN = Side(style="thin", color="BFC9CF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = Workbook()

# =====================================================================
# Sheet 1 -- Instructions
# =====================================================================
ins = wb.active
ins.title = "Instructions"
ins.sheet_view.showGridLines = False

rows = [
    ("The Empty CMDB - capability readiness matrix", "title"),
    ("IAAI-27 submission - practitioner input required", "sub"),
    ("", ""),
    ("What this is", "h"),
    ("Ten AIOps capability classes. Everything measurable from the three public ITSM "
     "instances is already filled in. Four columns need your professional judgement "
     "because they cannot be derived from data.", "p"),
    ("", ""),
    ("What you need to do", "h"),
    ("Open the 'Matrix' sheet and complete the four YELLOW columns (H, I, J, K) for "
     "each of the ten rows. Nothing else needs editing.", "p"),
    ("", ""),
    ("Column H - CSDM stage required", "h2"),
    ("At which CSDM stage does the data this capability needs become reliably "
     "available in a real client environment? Pick from the dropdown: Foundation, "
     "Crawl, Walk, Run, Fly, or Never. If you disagree with my 'required data' entry "
     "in column B, say so in column K.", "p"),
    ("", ""),
    ("Column I - Typical CI population at that stage", "h2"),
    ("In your experience, what percentage of incidents actually carry a populated "
     "configuration item at that CSDM stage? Enter a percentage. This is the single "
     "most valuable number you can give - the paper has a measured curve relating "
     "population rate to capability, but no way to anchor CSDM stages onto it. "
     "You are the anchor.", "p"),
    ("", ""),
    ("Column J - Client demand", "h2"),
    ("How often do clients actually ask for this capability? High, Medium, or Low. "
     "This decides which capabilities the paper foregrounds.", "p"),
    ("", ""),
    ("Column K - Your verdict or correction", "h2"),
    ("Free text. Correct my 'required data' if it is wrong, note caveats, or flag "
     "capabilities I have missed. Disagreement here is more useful than agreement.", "p"),
    ("", ""),
    ("What happens automatically", "h"),
    ("Column L is a formula. Once you enter a population rate in column I, it "
     "interpolates the modelled benefit retained from the measured degradation curve "
     "(Evidence sheet, rows 6-19). Do not edit column L.", "p"),
    ("", ""),
    ("Worked example - how a completed row should look", "h"),
]
r = 1
for text, kind in rows:
    c = ins.cell(row=r, column=1, value=text)
    if kind == "title":
        c.font = Font(ARIAL, 16, bold=True, color="0D6B6E")
    elif kind == "sub":
        c.font = Font(ARIAL, 11, italic=True, color="66747F")
    elif kind == "h":
        c.font = Font(ARIAL, 12, bold=True)
    elif kind == "h2":
        c.font = Font(ARIAL, 10.5, bold=True, color="0D6B6E")
    else:
        c.font = Font(ARIAL, 10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ins.row_dimensions[r].height = 42
    r += 1

ex_hdr = ["Capability", "CSDM stage (H)", "Typical CI pop. (I)",
          "Demand (J)", "Your verdict (K)"]
ex_val = ["Misrouting risk prediction", "Walk", 0.65, "High",
          "Agree. Needs Application Service mapping done, not just CI class "
          "population - a CI with no service relationship still routes badly."]
for j, h in enumerate(ex_hdr, start=1):
    c = ins.cell(row=r, column=j, value=h)
    c.font = Font(ARIAL, 9, bold=True, color="FFFFFF")
    c.fill = HDR_FILL
    c.alignment = CTR
    c.border = BOX
r += 1
for j, v in enumerate(ex_val, start=1):
    c = ins.cell(row=r, column=j, value=v)
    c.font = Font(ARIAL, 9)
    c.alignment = WRAP
    c.border = BOX
    if j in (2, 3, 4, 5):
        c.fill = INPUT_FILL
    if j == 3:
        c.number_format = "0%"
ins.row_dimensions[r].height = 46
r += 2
c = ins.cell(row=r, column=1,
             value="Legend:  YELLOW = you fill this in.   White = already derived from "
                   "the data, leave alone.   Pale teal = calculated by formula, do not edit.")
c.font = Font(ARIAL, 9.5, italic=True, color="3D4A54")

for col, w in zip("ABCDE", (46, 18, 20, 14, 52)):
    ins.column_dimensions[col].width = w

# =====================================================================
# Sheet 2 -- Matrix
# =====================================================================
mx = wb.create_sheet("Matrix")
mx.sheet_view.showGridLines = False

HEADERS = [
    ("Capability", 30),
    ("Required data element(s)  [derived]", 34),
    ("Data axis", 14),
    ("VolvoIT", 12),
    ("ServiceNow-IT", 13),
    ("Rabobank", 12),
    ("Measured in this study", 30),
    ("CSDM stage required", 16),
    ("Typical CI population at that stage", 16),
    ("Client demand", 12),
    ("Your verdict / correction", 40),
    ("Modelled benefit retained", 15),
]

DATA = [
    ("Misrouting risk prediction",
     "Component identity at creation (CI or equivalent intake field)",
     "identity", "product field (100%)", "cmdb_ci 0.2%", "CI Name 99.6%",
     "MEASURED. AUC 0.566 -> 0.743 with CI (+0.176, CI [+0.166,+0.188])"),
    ("Service-target breach prediction",
     "Component identity at creation",
     "identity", "not tested", "cmdb_ci 0.2%", "CI Name 99.6%",
     "MEASURED. Rabobank 0.653 -> 0.761 (+0.108). ServiceNow +0.002"),
    ("Automated assignment / routing",
     "Component identity + historical routing outcomes per component",
     "identity", "product + org:group", "cmdb_ci 0.2%", "CI Name + WBS",
     "Not directly measured; same input as misrouting prediction"),
    ("Priority / impact auto-classification",
     "Intake fields only (category, impact, urgency)",
     "intake", "available", "available", "available",
     "Intake fields are ~100% populated in all three instances"),
    ("Knowledge article recommendation",
     "Knowledge reference + symptom or category",
     "linkage", "absent", "knowledge flag only", "KM number 99.6%",
     "Not measured. Link present only in Rabobank"),
    ("Problem clustering / recurring incident detection",
     "Incident to problem reference",
     "linkage", "absent", "problem_id 1.5%", "# Related Incidents 2.6%",
     "Not measurable in any instance"),
    ("Change-incident correlation",
     "Incident to change reference",
     "linkage", "absent", "rfc 0.7%", "Related Change 1.2%",
     "Unavailable even at 99.6% CI maturity. 30,275 change records exist "
     "in the same export but are not linked"),
    ("CI-topology root cause analysis",
     "CI to CI dependency graph",
     "topology", "absent", "absent", "absent",
     "NO dependency edges in any of the three exports. Identity and "
     "classification are present; topology is not"),
    ("Blast-radius / impact estimation",
     "CI topology + change reference",
     "topology", "absent", "absent", "absent",
     "Requires both missing axes simultaneously"),
    ("Causal chain reconstruction",
     "Causing-incident reference",
     "linkage", "absent", "caused_by 0.0%", "absent",
     "Field exists in ServiceNow schema and is never populated"),
]

for j, (h, w) in enumerate(HEADERS, start=1):
    c = mx.cell(row=1, column=j, value=h)
    c.font = Font(ARIAL, 9, bold=True, color="FFFFFF")
    c.fill = HDR_FILL
    c.alignment = CTR
    c.border = BOX
    mx.column_dimensions[get_column_letter(j)].width = w
mx.row_dimensions[1].height = 42

for i, row in enumerate(DATA, start=2):
    for j, v in enumerate(row, start=1):
        c = mx.cell(row=i, column=j, value=v)
        c.font = Font(ARIAL, 9, bold=(j == 1))
        c.alignment = WRAP
        c.border = BOX
    for j in (8, 9, 10, 11):
        c = mx.cell(row=i, column=j)
        c.fill = INPUT_FILL
        c.border = BOX
        c.alignment = WRAP
    mx.cell(row=i, column=9).number_format = "0%"

    # column L -- linear interpolation into the measured degradation curve.
    # Guards force 0 < I < 1 before MATCH runs, so the bracketing row m and
    # its successor m+1 always exist and no #REF! is possible.  Non-numeric
    # or blank input returns "" rather than an error.
    B, C = "Evidence!$B$6:$B$19", "Evidence!$C$6:$C$19"
    m = f"MATCH(I{i},{B},1)"
    interp = (f"INDEX({C},{m})"
              f"+(I{i}-INDEX({B},{m}))"
              f"/(INDEX({B},{m}+1)-INDEX({B},{m}))"
              f"*(INDEX({C},{m}+1)-INDEX({C},{m}))")
    f = (f'=IF(NOT(ISNUMBER(I{i})),"",'
         f'IF(I{i}>=1,1,IF(I{i}<=0,0,{interp})))')
    c = mx.cell(row=i, column=12, value=f)
    c.fill = CALC_FILL
    c.font = Font(ARIAL, 9)
    c.number_format = "0.0%"
    c.alignment = CTR
    c.border = BOX
    mx.row_dimensions[i].height = 54

dv_stage = DataValidation(
    type="list",
    formula1='"Foundation,Crawl,Walk,Run,Fly,Never"',
    allow_blank=True, showDropDown=False)
dv_dem = DataValidation(type="list", formula1='"High,Medium,Low"',
                        allow_blank=True, showDropDown=False)
mx.add_data_validation(dv_stage)
mx.add_data_validation(dv_dem)
dv_stage.add(f"H2:H{len(DATA)+1}")
dv_dem.add(f"J2:J{len(DATA)+1}")

n = len(DATA) + 3
mx.cell(row=n, column=1,
        value="YELLOW columns (H-K) need your input. Column L calculates itself from "
              "column I. Everything else is measured - see the Evidence sheet."
        ).font = Font(ARIAL, 9, italic=True, color="3D4A54")
mx.freeze_panes = "B2"

# =====================================================================
# Sheet 3 -- Evidence
# =====================================================================
ev = wb.create_sheet("Evidence")
ev.sheet_view.showGridLines = False


def section(row, title):
    c = ev.cell(row=row, column=1, value=title)
    c.font = Font(ARIAL, 11, bold=True, color="0D6B6E")
    ev.cell(row=row, column=1).fill = SECT_FILL
    for j in range(2, 7):
        ev.cell(row=row, column=j).fill = SECT_FILL


def table(start, headers, data, pct_cols=(), dec_cols=()):
    for j, h in enumerate(headers, start=1):
        c = ev.cell(row=start, column=j, value=h)
        c.font = Font(ARIAL, 9, bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = CTR
        c.border = BOX
    for i, row in enumerate(data, start=start + 1):
        for j, v in enumerate(row, start=1):
            c = ev.cell(row=i, column=j, value=v)
            c.font = Font(ARIAL, 9)
            c.border = BOX
            if j in pct_cols:
                c.number_format = "0.0%"
            if j in dec_cols:
                c.number_format = "0.000"
    return start + len(data) + 2


ev.cell(row=1, column=1, value="Measured evidence").font = Font(ARIAL, 14, bold=True)
ev.cell(row=2, column=1,
        value="79,281 incidents - 3 organisations - 3 platforms - all public data. "
              "Generated by scripts/e1..e8. Do not edit."
        ).font = Font(ARIAL, 9.5, italic=True, color="66747F")

section(4, "Degradation curve - Rabobank CI fields blanked in silico (drives column L)")
DEG = [(0.000, 0.000), (0.002, 0.000), (0.010, 0.003), (0.020, 0.017),
       (0.050, 0.040), (0.100, 0.089), (0.200, 0.215), (0.300, 0.330),
       (0.400, 0.449), (0.500, 0.560), (0.600, 0.669), (0.750, 0.813),
       (0.900, 0.938), (1.000, 1.000)]
# NOTE: the Matrix column-L formula reads rate from column B and benefit from
# column C, so the label column A must stay in place.  Do not remove it.
nxt = table(5, ["level", "CI population rate", "benefit retained", "", "", ""],
            [(f"L{k:02d}", a, b, "", "", "") for k, (a, b) in enumerate(DEG, 1)],
            pct_cols=(2, 3))

section(nxt, "Field population by semantic class (E1)")
POP = [("intake", 1.000, 0.997, 0.996), ("descriptive", 1.000, 0.941, 0.996),
       ("workflow", 0.968, 0.971, 0.996), ("outcome", 1.000, 0.998, 0.981),
       ("configuration", None, 0.002, 0.996), ("relational", None, 0.006, 0.498)]
POP = [(a, "absent" if b is None else b, c, d, "", "") for a, b, c, d in POP]
nxt = table(nxt + 1, ["field class", "VolvoIT", "ServiceNow-IT", "Rabobank", "", ""],
            POP, pct_cols=(2, 3, 4))

section(nxt, "Misrouting prediction, creation-time features, temporal split (E2)")
E2 = [("ServiceNow-IT", 0.699, 0.500, 0.699, "-0.000", "[-0.002,+0.002]"),
      ("VolvoIT", 0.868, None, None, "n/a", "no CMDB captured"),
      ("Rabobank", 0.566, 0.742, 0.743, "+0.176", "[+0.166,+0.188]")]
E2 = [(a, b, c if c else "-", d if d else "-", e, f) for a, b, c, d, e, f in E2]
nxt = table(nxt + 1,
            ["organisation", "intake", "config-only", "intake+config",
             "delta", "95% CI"], E2, dec_cols=(2, 3, 4))

section(nxt, "Single-field predictive power - component vs person identity (E5)")
E5 = [("product", "VolvoIT", 615, 0.743, "component", ""),
      ("CI Name (aff)", "Rabobank", 2633, 0.731, "component", ""),
      ("Service Component WBS", "Rabobank", 259, 0.712, "component", ""),
      ("category", "ServiceNow-IT", 43, 0.619, "coarse service", ""),
      ("subcategory", "ServiceNow-IT", 210, 0.615, "coarse service", ""),
      ("opened_by", "ServiceNow-IT", 151, 0.539, "person", ""),
      ("caller_id", "ServiceNow-IT", 4526, 0.518, "person", ""),
      ("cmdb_ci", "ServiceNow-IT", 40, 0.500, "component (0.2% populated)", "")]
nxt = table(nxt + 1, ["field", "organisation", "cardinality", "single-field AUC",
                      "identifies", ""], E5, dec_cols=(4,))

section(nxt, "Cost of misrouting (E4)")
COST = [("ServiceNow-IT", "SLA breach, 0 reassignments", 0.216, "", "", ""),
        ("ServiceNow-IT", "SLA breach, 5+ reassignments", 0.857, "", "", ""),
        ("Rabobank", "median handling, routed correctly", 1.7, "hours", "", ""),
        ("Rabobank", "median handling, later reassigned", 15.3, "hours", "", ""),
        ("Rabobank", "excess per misroute", 13.6, "hours", "", ""),
        ("Rabobank", "review queue precision at 5% capacity, no CMDB", 0.531, "", "", ""),
        ("Rabobank", "review queue precision at 5% capacity, with CMDB", 0.913, "", "", "")]
nxt = table(nxt + 1, ["organisation", "measure", "value", "unit", "", ""], COST)

section(nxt, "Robustness (E8) - Rabobank configuration delta")
ROB = [("across 3 learners x 4 temporal splits", "min", 0.176, "", "", ""),
       ("", "median", 0.190, "", "", ""),
       ("", "max", 0.210, "", "", ""),
       ("", "positive in", "12 of 12", "", "", "")]
table(nxt + 1, ["scope", "statistic", "value", "", "", ""], ROB, dec_cols=(3,))

for col, w in zip("ABCDEF", (34, 34, 16, 18, 22, 20)):
    ev.column_dimensions[col].width = w

wb.save(OUT)
print(f"written: {OUT}")
